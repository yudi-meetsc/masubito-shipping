import io
import zipfile
from collections import defaultdict
from typing import List, Dict, Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from app.utils.excel_helpers import set_str

_FILLS = {
    "main": PatternFill("solid", fgColor="1F4E79"),
    "cooler": PatternFill("solid", fgColor="375623"),
    "yupacket": PatternFill("solid", fgColor="7B2C2C"),
}
_HEADER_FONT = Font(bold=True, color="FFFFFF")

MAIN_COOLER_HEADERS = [
    "お届け先コード取得区分", "お届け先コード",
    "お届け先電話番号", "お届け先郵便番号",
    "お届け先住所１", "お届け先住所２", "お届け先住所３",
    "お届け先名称１", "お届け先名称２",
    "お客様管理番号", "お客様コード",
    "部署ご担当者コード取得区分", "部署ご担当者コード", "部署ご担当者名称",
    "荷送人電話番号",
    "ご依頼主コード取得区分", "ご依頼主コード", "ご依頼主電話番号",
    "ご依頼主郵便番号", "ご依頼主住所１", "ご依頼主住所２",
    "ご依頼主名称１", "ご依頼主名称２",
    "荷姿",
    "品名１", "品名２", "品名３", "品名４", "品名５",
    "荷札荷姿",
    "荷札品名１", "荷札品名２", "荷札品名３", "荷札品名４", "荷札品名５",
    "荷札品名６", "荷札品名７", "荷札品名８", "荷札品名９", "荷札品名１０", "荷札品名１１",
    "出荷個数", "スピード指定", "クール便指定",
    "配達日", "配達指定時間帯", "配達指定時間（時分）",
    "代引金額", "消費税", "決済種別", "保険金額",
    "指定シール１", "指定シール２", "指定シール３",
    "営業所受取", "SRC区分", "営業所受取営業所コード",
    "元着区分", "メールアドレス", "ご不在時連絡先",
    "出荷日", "お問い合せ送り状No.", "出荷場印字区分", "集約解除指定",
    "編集０１", "編集０２", "編集０３", "編集０４", "編集０５",
    "編集０６", "編集０７", "編集０８", "編集０９", "編集１０",
]

YUPACKET_HEADERS = [
    "お届け先郵便番号",
    "お届け先住所1", "お届け先住所2", "お届け先住所3",
    "お届け先名称１", "お届け先名称２",
    "お届け先電話番号",
    "品名", "厚さ", "荷送人名", "記事欄",
    "管理番号",
]


def _prep(row: Dict[str, Any]):
    """Return (addr1, addr2, addr3, name1, name2, tel, postal)."""
    addr_all = (
        str(row.get("届け先都道府県", ""))
        + str(row.get("届け先住所１", ""))
        + str(row.get("届け先住所２", ""))
    ).replace("−", "-")
    addr1 = addr_all[:16]
    addr2 = addr_all[16:32]
    addr3 = addr_all[32:48]

    name = str(row.get("届け先氏名", ""))
    name1 = name[:16]
    name2 = name[16:] if len(name) > 16 else ""

    tel = str(row.get("届け先ＴＥＬ", "")).replace("−", "-")
    postal = str(row.get("届け先郵便番号", "")).replace("−", "-")
    return addr1, addr2, addr3, name1, name2, tel, postal


def _write_header(ws, headers, fill):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.font = _HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


def _write_main_cooler_row(ws, r_idx: int, row: Dict[str, Any], client_code: str) -> None:
    addr1, addr2, addr3, name1, name2, tel, postal = _prep(row)
    mgmt = row.get("管理番号", "")
    time_code = row.get("配送時間帯コード", "")
    delivery_date = row.get("配送希望日", "")

    # (value, force_string)  — 74 columns matching MAIN_COOLER_HEADERS order
    values = [
        ("", False),             # お届け先コード取得区分
        ("", False),             # お届け先コード
        (tel, True),             # お届け先電話番号
        (postal, True),          # お届け先郵便番号
        (addr1, False),          # お届け先住所１
        (addr2, True),           # お届け先住所２ (date-safe)
        (addr3, True),           # お届け先住所３ (date-safe)
        (name1, False),          # お届け先名称１
        (name2, False),          # お届け先名称２
        ("", False),             # お客様管理番号
        (client_code, True),     # お客様コード
        ("", False),             # 部署ご担当者コード取得区分
        ("", False),             # 部署ご担当者コード
        ("", False),             # 部署ご担当者名称
        ("", False),             # 荷送人電話番号
        ("", False),             # ご依頼主コード取得区分
        ("", False),             # ご依頼主コード
        ("", False),             # ご依頼主電話番号
        ("", False),             # ご依頼主郵便番号
        ("", False),             # ご依頼主住所１
        ("", False),             # ご依頼主住所２
        ("", False),             # ご依頼主名称１
        ("", False),             # ご依頼主名称２
        ("", False),             # 荷姿
        (mgmt, True),            # 品名１ (管理番号8桁string)
        ("釣具", False),         # 品名２
        ("", False),             # 品名３
        ("", False),             # 品名４
        ("", False),             # 品名５
        ("", False),             # 荷札荷姿
        ("", False),             # 荷札品名１
        ("", False),             # 荷札品名２
        ("", False),             # 荷札品名３
        ("", False),             # 荷札品名４
        ("", False),             # 荷札品名５
        ("", False),             # 荷札品名６
        ("", False),             # 荷札品名７
        ("", False),             # 荷札品名８
        ("", False),             # 荷札品名９
        ("", False),             # 荷札品名１０
        ("", False),             # 荷札品名１１
        ("", False),             # 出荷個数
        ("", False),             # スピード指定
        ("", False),             # クール便指定
        (delivery_date, False),  # 配達日
        (time_code, True),       # 配達指定時間帯
        ("", False),             # 配達指定時間（時分）
        ("", False),             # 代引金額
        ("", False),             # 消費税
        ("", False),             # 決済種別
        ("", False),             # 保険金額
        ("", False),             # 指定シール１
        ("", False),             # 指定シール２
        ("", False),             # 指定シール３
        ("", False),             # 営業所受取
        ("", False),             # SRC区分
        ("", False),             # 営業所受取営業所コード
        ("", False),             # 元着区分
        ("", False),             # メールアドレス
        ("", False),             # ご不在時連絡先
        ("", False),             # 出荷日
        ("", False),             # お問い合せ送り状No.
        ("", False),             # 出荷場印字区分
        ("", False),             # 集約解除指定
        ("", False),             # 編集０１
        ("", False),             # 編集０２
        ("", False),             # 編集０３
        ("", False),             # 編集０４
        ("", False),             # 編集０５
        ("", False),             # 編集０６
        ("", False),             # 編集０７
        ("", False),             # 編集０８
        ("", False),             # 編集０９
        ("", False),             # 編集１０
    ]

    for col_idx, (val, is_str) in enumerate(values, 1):
        cell = ws.cell(r_idx, col_idx)
        if is_str:
            set_str(cell, val)
        else:
            cell.value = val


def _write_yupacket_row(ws, r_idx: int, row: Dict[str, Any]) -> None:
    addr1, addr2, addr3, name1, name2, tel, postal = _prep(row)
    mgmt = row.get("管理番号", "")

    values = [
        (postal, True),          # お届け先郵便番号
        (addr1, False),          # お届け先住所1
        (addr2, True),           # お届け先住所2 (date-safe)
        (addr3, True),           # お届け先住所3 (date-safe)
        (name1, False),          # お届け先名称１
        (name2, False),          # お届け先名称２
        (tel, True),             # お届け先電話番号
        ("釣具", False),         # 品名
        (3, False),              # 厚さ (number)
        ("ますびと商店", False), # 荷送人名
        ("", False),             # 記事欄
        (mgmt, True),            # 管理番号
    ]

    for col_idx, (val, is_str) in enumerate(values, 1):
        cell = ws.cell(r_idx, col_idx)
        if is_str:
            set_str(cell, val)
        else:
            cell.value = val


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)


def _make_zip(label: str, ts: str, wb: openpyxl.Workbook) -> bytes:
    excel_buf = io.BytesIO()
    wb.save(excel_buf)
    filename_base = f"{ts}_注文データ_{label}"
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"{filename_base}.xlsx", excel_buf.getvalue())
    return zip_buf.getvalue()


def build(sakai_24: List[Dict[str, Any]], sakai_3: List[Dict[str, Any]], ts: str) -> Dict[str, bytes]:
    """Return dict with keys 'main', 'cooler', 'yupacket' → ZIP bytes (only if rows exist)."""
    result: Dict[str, bytes] = {}

    # Sort by 管理番号 ascending (matches GAS sort before grouping)
    sakai_24_sorted = sorted(sakai_24, key=lambda r: r.get("管理番号", ""))

    # Group by 管理番号, preserving sorted order
    groups: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    order: List[str] = []
    for row in sakai_24_sorted:
        key = row.get("管理番号", "")
        if key not in groups:
            order.append(key)
        groups[key].append(row)

    main_rows: List[Dict[str, Any]] = []
    yupacket_rows: List[Dict[str, Any]] = []

    for key in order:
        g = groups[key]
        shelves = {r["_shelf"] for r in g}
        if shelves == {"4"}:
            yupacket_rows.append(g[0])   # all shelf-4 → ゆうパケット (first row only)
        else:
            main_rows.append(g[0])       # all-2 or mixed 2+4 → メイン (first row only)

    # ---- 堺メイン ----
    if main_rows:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "堺メイン"
        _write_header(ws, MAIN_COOLER_HEADERS, _FILLS["main"])
        for r_idx, row in enumerate(main_rows, 2):
            _write_main_cooler_row(ws, r_idx, row, "145204860071")
        _auto_width(ws)
        result["main"] = _make_zip("堺メイン", ts, wb)

    # ---- 堺クーラー ----
    if sakai_3:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "堺クーラー"
        _write_header(ws, MAIN_COOLER_HEADERS, _FILLS["cooler"])
        for r_idx, row in enumerate(sakai_3, 2):
            _write_main_cooler_row(ws, r_idx, row, "14520486009")
        _auto_width(ws)
        result["cooler"] = _make_zip("堺クーラー", ts, wb)

    # ---- 堺ゆうパケット ----
    if yupacket_rows:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "堺ゆうパケット"
        _write_header(ws, YUPACKET_HEADERS, _FILLS["yupacket"])
        for r_idx, row in enumerate(yupacket_rows, 2):
            _write_yupacket_row(ws, r_idx, row)
        _auto_width(ws)
        result["yupacket"] = _make_zip("堺ゆうパケット", ts, wb)

    return result
