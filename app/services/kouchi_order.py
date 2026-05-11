import io
import zipfile
from typing import List, Dict, Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from app.utils.excel_helpers import set_str

HEADERS = [
    "店舗名", "注文管理番号", "JAN（商品コード）", "商品名", "数量",
    "税込販売単価", "備考", "電話番号", "郵便番号", "住所1", "住所2", "住所3",
    "お客様名", "配達日", "配達指定時間帯", "サイズ", "備考",
]

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def build(rows: List[Dict[str, Any]], ts: str) -> bytes:
    """Return ZIP bytes containing the Kouchi order Excel (1 row per CSV row)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "高知注文データ"

    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(1, col, h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for r_idx, row in enumerate(rows, 2):
        col = 1
        ws.cell(r_idx, col, row.get("店舗名", "")); col += 1
        set_str(ws.cell(r_idx, col), row.get("管理番号", "")); col += 1
        set_str(ws.cell(r_idx, col), row["_product_key"]); col += 1
        ws.cell(r_idx, col, row.get("標準商品名", "")); col += 1
        qty = row.get("数量", "")
        ws.cell(r_idx, col, int(qty) if qty.isdigit() else qty); col += 1
        price = row.get("単価", "")
        try:
            ws.cell(r_idx, col, float(price) if price else "")
        except ValueError:
            ws.cell(r_idx, col, price)
        col += 1
        ws.cell(r_idx, col, row.get("配送備考", "")); col += 1
        set_str(ws.cell(r_idx, col), row.get("届け先ＴＥＬ", "")); col += 1
        set_str(ws.cell(r_idx, col), row.get("届け先郵便番号", "")); col += 1
        ws.cell(r_idx, col, row.get("届け先都道府県", "")); col += 1
        ws.cell(r_idx, col, row.get("届け先住所１", "")); col += 1
        set_str(ws.cell(r_idx, col), row.get("届け先住所２", "")); col += 1
        ws.cell(r_idx, col, row.get("届け先氏名", "")); col += 1
        ws.cell(r_idx, col, row.get("配送希望日", "")); col += 1
        set_str(ws.cell(r_idx, col), row.get("配送時間帯コード", "")); col += 1
        ws.cell(r_idx, col, row.get("商品備考", "")); col += 1
        ws.cell(r_idx, col, row.get("備考", "")); col += 1

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    excel_buf = io.BytesIO()
    wb.save(excel_buf)

    filename_base = f"{ts}_注文データ_高知"
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"{filename_base}.xlsx", excel_buf.getvalue())
    return zip_buf.getvalue()
