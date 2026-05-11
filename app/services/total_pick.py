import io
import zipfile
from typing import List, Dict, Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from app.utils.excel_helpers import set_str

_HEADER_FILL = PatternFill("solid", fgColor="375623")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADERS = ["JAN（商品コード）", "商品名", "数量", "ルアー"]


def build(rows: List[Dict[str, Any]], ts: str) -> bytes:
    """Aggregate all non-kouchi rows by product key (JAN priority) and build total pick Excel."""
    aggregated: Dict[str, Dict[str, Any]] = {}
    order: List[str] = []

    for row in rows:
        key = row["_product_key"]
        if not key:
            continue

        qty_str = row.get("数量", "0")
        qty = int(qty_str) if qty_str.isdigit() else 0

        if key not in aggregated:
            name = (
                row.get("標準商品名", "")
                + row.get("属性１名", "")
                + row.get("属性２名", "")
            )
            aggregated[key] = {
                "name": name,
                "quantity": 0,
                "tanaban": row.get("_shelf", ""),  # first occurrence's shelf for ルアー check
            }
            order.append(key)
        aggregated[key]["quantity"] += qty

    # Sort by key (JAN/code) ascending — matches GAS sortSheetExcludingHeader
    sorted_keys = sorted(order)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "トータルピック表"

    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(1, col, h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    total_sum = 0
    lure_sum = 0

    for r_idx, key in enumerate(sorted_keys, 2):
        item = aggregated[key]
        qty = item["quantity"]
        total_sum += qty
        is_lure = item["tanaban"] == "4"
        if is_lure:
            lure_sum += qty

        set_str(ws.cell(r_idx, 1), key)
        ws.cell(r_idx, 2, item["name"])
        ws.cell(r_idx, 3, qty)
        ws.cell(r_idx, 4, qty if is_lure else "")

    # Total row: only C (sum) and D (lure sum), A and B left blank
    total_row = len(sorted_keys) + 2
    ws.cell(total_row, 3, total_sum)
    ws.cell(total_row, 4, lure_sum)

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    excel_buf = io.BytesIO()
    wb.save(excel_buf)

    filename_base = f"{ts}_トータルピック表"
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"{filename_base}.xlsx", excel_buf.getvalue())
    return zip_buf.getvalue()
